"""
Class to hold original question store (from file),
and results of duplication detection.
"""

import logging
import os
from typing import List

import dotsi  # type: ignore
import openpyxl
import spacy
from spacy.lang.en.stop_words import STOP_WORDS
from spacy.language import Language
import spacy_universal_sentence_encoder

import doubles.progress as prog

# Get looger for this application
log = logging.getLogger(__name__)

# Load the spaCy model.
# Had to download separately, and doing via pyproject.toml file didn't seem to work.
# poetry run python -m spacy download en_core_web_md
# Didn't need to do this for the universal sentence encoder.
nlp = spacy_universal_sentence_encoder.load_model("en_use_lg")


class Question:
    """
    Question class.
    """

    def __init__(self, lid: int, question_text: str, answer: bool) -> None:
        """
        Question initialisation.
        Args:
            lid:            Legacy Id.
            question_text:  Question text (including ?).
            answer:         Expected response True=yes, False=no.
        Returns:
            Returns         None.
        """

        # Id of the source question.
        self.lid = lid

        # Raw question text.
        self.original_question = question_text
        # Normalise the string and convert to tokens.
        self.tokens = self.original_question.lower().split()
        # Remove punctuation.
        self.tokens = [token.strip("?,.!") for token in self.tokens]
        # Remove common words that bias matching (stop words).
        self.tokens = [token for token in self.tokens if token not in list(STOP_WORDS)]
        # Reform tokens to processing string.
        self.question = " ".join(self.tokens)

        # Expected answer, True=yes, False=no.
        self.answer = answer

        # Whether or not question has been the reference question yet.
        self.reference = False
        # Question statuses.
        self.unique = True
        self.duplicate = False
        # List of duplicates to this question.
        self.duplicates: List[Question] = []
        # The question this is a duplicate of (only the first).
        self.duplicate_of = 0


class Question_Store:
    """
    Class for the store of questions.
    """

    def __init__(self, ifile: str, settings: dotsi.Dict) -> None:
        """
        Question_Store initialisation.
        Args:
            ifile:      Question file (Excel) containing questions.
            settings:   Applications settings.
        Returns:
            Returns None.
        """

        log.info("Initialising question store.")

        # Settings file.
        self.settings = settings

        # Initialise store.
        self.store = []
        self.num_q = 0
        self.status = self.settings.status.ST_NOQ

        # Load the workbook
        workbook = openpyxl.load_workbook(filename=ifile)

        # Select the active sheet
        sheet = workbook.active

        # Iterate through rows in the worksheet and extract question details.
        for row_num in range(2, sheet.max_row + 1):
            q_text = sheet.cell(row=row_num, column=1).value
            q_answer = bool(sheet.cell(row=row_num, column=2).value)

            self.store.append(Question(row_num - 1, q_text, q_answer))
            self.num_q += 1

        log.info(f"Number of questions read from file: {self.num_q}")

    def process(self, progress) -> None:
        """
        Process questions.
        Look for duplicates questions.
        The rest will be unique questions.
        Args:
            progress:       True if progress bar displayed.
        Returns:
            Returns None.
        """

        log.info("Processing questions...")

        # Number of duplicate questions detected, and duplicates with errors.
        self.num_duplicates = 0

        # If only one question, nothing to test as questions unique.
        if self.num_q == 1:
            self.status = self.settings.status.ST_NOQ
            log.warning(f"Status: {self.status}")
            return

        # Set up the progress bar.
        if progress is True:
            pb = prog.CLI_PROGRESS(self.settings, "Analysing")

        # Else, process questions.
        # Progress increment for reference question (percent).
        pps = 100 / self.num_q

        for idx, q in enumerate(self.store[0:]):
            if progress is True:
                pb.show_progress(int((idx + 1) * pps), q.lid)

            # Need to find the first reference question.
            # That is question that hasn't been compared against yet,
            # or is not already a duplicate question.
            if q.reference is True or q.duplicate is True:
                continue
            else:
                # Mark this question as being a reference question.
                q.reference = True
                log.debug(f"Ref. question, id: {q.lid}, Text: {q.original_question}, Tokens: {q.question}")

                # Need to find next question that isn't a duplicate.
                for idx2, q2 in enumerate(self.store[idx + 1 :]):
                    if q2.duplicate is True:
                        if progress is True:
                            pb.show_progress(int((idx + 1) * pps), q.lid)
                        continue
                    else:
                        # Have 2 questions to compare now.
                        # Check similarity.
                        similarity = get_similarity(q.question, q2.question)
                        if similarity > self.settings.scores.SS_MATCH:
                            # Similarity score close enough for a match.
                            q2.unique = False
                            q2.duplicate = True
                            q.duplicates.append(q2)
                            self.num_duplicates += 1
                            q2.duplicate_of = q.lid
                            log.debug(
                                f"DUPLICATE, id: {q2.lid}, Text: {q2.original_question}, Tokens: {q2.question}, similarity: {similarity :.3f}"
                            )
                        else:
                            # Not a match.
                            log.debug(
                                f"Checked question, id: {q2.lid}, Text: {q2.original_question}, Tokens: {q2.question}, similarity: {similarity :.3f}"
                            )

    def results(self) -> None:
        """
        Output results of processing.
        Args:
        Returns:
            Returns None.
        """

        log.info("Generating results...")

        # Report statistcis on level of duplication detection.
        print("\n")
        print("*" * 80)
        print(f"Questions in orginal file           : {self.num_q}")
        print(f"Duplicate questions found           : {self.num_duplicates}")
        print(f"Duplicate questions (%)             : {self.num_duplicates / self.num_q * 100 :.1f} %")
        print("*" * 80)

    def export(self, ofile) -> None:
        """
        Export results of processing to out file.
        Args:
            ofile:      Output file.
        Returns:
            Returns None.
        """

        log.info("Export results to file...")

        # Generate the output file with duplicates removed.
        work_book = openpyxl.Workbook()
        sheet1 = work_book.active
        sheet1.title = "Unique_Questions"

        # Write title row.
        sheet1["A1"].value = "Id"
        sheet1["B1"].value = "Question"
        sheet1["C1"].value = "Answer"

        # Export questions with duplicates removed.
        ex_row = 2
        for idx, q in enumerate(self.store[0:]):
            if q.duplicate is False:
                sheet1.cell(row=ex_row, column=1).value = q.lid
                sheet1.cell(row=ex_row, column=2).value = q.original_question
                sheet1.cell(row=ex_row, column=3).value = "Yes" if q.answer == 1 else "No"
                ex_row += 1

        # Export duplicates to a separate sheet.
        work_book.create_sheet(title="Duplicates")
        sheet2 = work_book["Duplicates"]

        # Write title row.
        sheet2["A1"].value = "Id"
        sheet2["B1"].value = "Question"
        sheet2["C1"].value = "Answer"
        sheet2["D1"].value = "Duplicate of"

        # Export duplicate questions.
        ex_row = 2
        for idx, q in enumerate(self.store[0:]):
            if q.duplicate is True:
                sheet2.cell(row=ex_row, column=1).value = q.lid
                sheet2.cell(row=ex_row, column=2).value = q.original_question
                sheet2.cell(row=ex_row, column=3).value = "Yes" if q.answer == 1 else "No"
                sheet2.cell(row=ex_row, column=4).value = q.duplicate_of
                ex_row += 1

        # Save the workbook.
        work_book.save(ofile)


def get_similarity(q1: str, q2: str) -> float:
    """
    Function to use natural language processor to compare 2 questions for similarity.
        Args:
            q1:         Reference question.
            q2:         Question to compare against reference.
        Returns:
            float:      Similarity between questions, 0 to 1.
    """

    # Process the input questions.
    proc_q1 = nlp(q1)
    proc_q2 = nlp(q2)

    # Determine the similarity score.
    return proc_q1.similarity(proc_q2)
