import dotsi
import openpyxl

from doubles.app_settings import load
from doubles.question_store import Question_Store

def test_doubles():
    """
    Case with unique and duplicate questiona.
    """

    # Create a short input questions file.
    work_book = openpyxl.Workbook()
    sheet1 = work_book.active
    sheet1.title = "Input"

    # Write title row.
    sheet1["A1"].value = "Question"
    sheet1["B1"].value = "Answer"

    # Write the test questions.
    sheet1.cell(row=2, column=1).value = "Do you have the right footware?"
    sheet1.cell(row=2, column=2).value = "yes"
    sheet1.cell(row=3, column=1).value = "Do you have the correct footware?"
    sheet1.cell(row=3, column=2).value = "yes"
    sheet1.cell(row=4, column=1).value = "Is your shirt nice and bright?"
    sheet1.cell(row=4, column=2).value = "yes"

        # Save the workbook.
    work_book.save("./tests/test-input.xlsx")

    settings = dotsi.Dict(load("./doubles/settings.yaml"))
    questions = Question_Store("./tests/test-input.xlsx", settings)
    questions.process(None)
    questions.results()

    # Check the expected results,
    # i.e. 2nd question a duplicate of the first,
    # The third question unique.
    assert(questions.num_q == 3)
    assert(questions.store[0].unique == True)
    assert(questions.store[1].duplicate == True)
    assert(questions.store[1].duplicate_of == 1)
    assert(questions.store[2].unique == True)
